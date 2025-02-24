import requests
import csv
import time
import os
from openpyxl import Workbook, load_workbook

# URL da API
url = "http://transparencia.al.gov.br/despesa/json-despesa-avancada-filtro/"

# Parâmetros da consulta, incluindo todas as colunas e sem filtros específicos
params = {
    'data_registro_dti_': '01/01/2021',
    'data_registro_dtf_': '31/12/2021',
    'visualizar': [
        'orgao_descricao',
         'descricao_ug', 'ug', 'pt_funcao_id__descricao_funcao',
        'sub_funcao_id__descricao_sub_funcao', 'programa_id__programa_descricao',
        'projeto_atividade_id__projeto_descricao', 'fonte_mae_id__descricao_fonte_mae', 'codigo_favorecido',
        'fonte_id__descricao_fonte', 'descricao_subtitulo', 'descricao_natureza2',
        'descricao_natureza3', 'descricao_natureza5', 'descricao_natureza6',
        'descricao_natureza', 'ano', 'mes', 'nome_favorecido',
        'pt_funcao_id', 'fonte_mae_id', 'descricao_natureza3', 'descricao_natureza2',
        'descricao_natureza5', 'descricao_natureza6', 'programa_id', 'projeto_atividade_id'
    ],
    'valor': ['empenhado', 'liquidado', 'pago'],
    'limit': 1000000,
    'offset': 0
}

# Define os nomes dos arquivos
csv_filename = 'dados.csv'
excel_filename = 'dadosx.xlsx'

# Verifica se os arquivos já existem
csv_exists = os.path.isfile(csv_filename)
excel_exists = os.path.isfile(excel_filename)

# Loop para tentar obter os dados da API
while True:
    try:
        # Faz a requisição GET
        response = requests.get(url, params=params)

        # Verifica se a requisição foi bem-sucedida
        if response.status_code == 200:
            # Converte o resultado para JSON
            data = response.json()

            # Extrai os dados de "rows"
            rows = data['rows']

            # Caso as linhas existam, pega as chaves da primeira linha para usar como cabeçalhos
            if rows:
                # Obtém as chaves dos dados (nomes das colunas)
                headers = list(rows[0].keys())

                # Salva no CSV
                with open(csv_filename, mode='a', newline='', encoding='utf-8') as csv_file:
                    csv_writer = csv.writer(csv_file)
                    # Se o arquivo CSV ainda não existir, escreve o cabeçalho
                    if not csv_exists:
                        csv_writer.writerow(headers)
                        csv_exists = True
                    # Adiciona as novas linhas no CSV
                    for row in rows:
                        csv_writer.writerow([row.get(header, '') for header in headers])

                # Salva no Excel
                if excel_exists:
                    # Carrega o arquivo Excel existente
                    workbook = load_workbook(excel_filename)
                    sheet = workbook.active
                else:
                    # Cria um novo arquivo Excel
                    workbook = Workbook()
                    sheet = workbook.active
                    # Escreve o cabeçalho no Excel
                    sheet.append(headers)
                    excel_exists = True

                # Adiciona as novas linhas no Excel
                for row in rows:
                    sheet.append([row.get(header, '') for header in headers])

                # Salva o arquivo Excel
                workbook.save(excel_filename)

            print(f"Novos dados anexados com sucesso em '{csv_filename}' e '{excel_filename}'!")
            break  # Sai do loop após sucesso

        else:
            print(f"Erro: {response.status_code}. Tentando novamente em 1 minuto...")
            time.sleep(60)  # Aguarda 1 minuto antes de tentar novamente

    except Exception as e:
        print(f"Erro crítico: {e}. Tentando novamente em 1 minuto...")
        time.sleep(60)  # Aguarda 1 minuto antes de tentar novamente
