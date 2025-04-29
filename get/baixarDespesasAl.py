import threading
import requests
import csv
import time
import os
import datetime
from openpyxl import Workbook, load_workbook
import schedule

# URL da API
URL =  "http://transparencia.al.gov.br/despesa/json-despesa-avancada-filtro/"


# Define os nomes dos arquivos
CSV_FILENAME = 'dados aguia/recursos recebidos al/dados.csv'
EXCEL_FILENAME = 'dados aguia/recursos recebidos al/dadosx.xlsx'

def ultimo_dia_do_mes(ano, mes):
    """Retorna o último dia do mês para o ano e mês especificados."""
    if mes == 12:
        return 31
    return (datetime.date(ano, mes + 1, 1) - datetime.timedelta(days=1)).day

def baixar_dados(mes, ano):
    """
    Baixa os dados do mês e ano especificados e adiciona aos arquivos existentes.
    """
    params = {
        'data_registro_dti_': f'01/{mes:02d}/{ano}',
        'data_registro_dtf_': f'{ultimo_dia_do_mes(ano, mes)}/{mes:02d}/{ano}',
        'visualizar': [
            'orgao_descricao', 'descricao_ug', 'ug', 'pt_funcao_id__descricao_funcao',
            'sub_funcao_id__descricao_sub_funcao', 'programa_id__programa_descricao',
            'projeto_atividade_id__projeto_descricao', 'fonte_mae_id__descricao_fonte_mae', 'codigo_favorecido',
            'fonte_id__descricao_fonte', 'descricao_subtitulo', 'descricao_natureza2',
            'descricao_natureza3', 'descricao_natureza5', 'descricao_natureza6', 'descricao_natureza',
            'ano', 'mes', 'nome_favorecido', 'pt_funcao_id', 'fonte_mae_id', 'descricao_natureza3',
            'descricao_natureza2', 'descricao_natureza5', 'descricao_natureza6', 'programa_id', 'projeto_atividade_id'
        ],
        'valor': ['empenhado', 'liquidado', 'pago'],
        'limit': 1000000,
        'offset': 0
    }

    try:
        response = requests.get(URL, params=params)
        if response.status_code == 200:
            data = response.json()
            rows = data.get('rows', [])
            if rows:
                salvar_dados(rows)
                print(f"Dados de {mes}/{ano} baixados e adicionados com sucesso.")
            else:
                print(f"Nenhum dado encontrado para {mes}/{ano}.")
        else:
            print(f"Erro ao baixar os dados: {response.status_code}")
    except Exception as e:
        print(f"Erro crítico ao baixar dados: {e}")

def salvar_dados(rows):
    """Salva os dados no arquivo CSV e Excel."""
    headers = list(rows[0].keys()) if rows else []
    
    # Salva no CSV
    csv_exists = os.path.isfile(CSV_FILENAME)
    with open(CSV_FILENAME, mode='a', newline='', encoding='utf-8') as csv_file:
        csv_writer = csv.writer(csv_file)
        if not csv_exists:
            csv_writer.writerow(headers)
        for row in rows:
            csv_writer.writerow([row.get(header, '') for header in headers])
    
    # Salva no Excel
    if os.path.isfile(EXCEL_FILENAME):
        workbook = load_workbook(EXCEL_FILENAME)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
    
    for row in rows:
        sheet.append([row.get(header, '') for header in headers])
    
    workbook.save(EXCEL_FILENAME)

def verificar_e_executar():
    """Verifica se é o dia e executa a coleta de dados do mês anterior."""
    hoje = datetime.datetime.today()
    if hoje.day == 1:
        mes_anterior = 12 if hoje.month == 1 else hoje.month - 1
        ano_anterior = hoje.year - 1 if hoje.month == 1 else hoje.year
        print(f"📅 Executando coleta para {mes_anterior:02d}/{ano_anterior}.")
        baixar_dados(mes_anterior, ano_anterior)
        print("🛑 Coleta concluída. Encerrando o programa.")
        os._exit(0)  # Encerra o programa

def agendar_execucao():
    """Agendar a execução da coleta de dados mensalmente"""
    schedule.every().day.at("08:55").do(verificar_e_executar)
    while True:
        schedule.run_pending()
        time.sleep(60)

if __name__ == "__main__":
    threading.Thread(target=agendar_execucao, daemon=True).start()
    while True:
        time.sleep(1)
